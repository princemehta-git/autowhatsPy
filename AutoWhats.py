import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter import ttk
# from threading import Thread

#  pip install -U pypiwin32

root = tk.Tk()

# setting the windows size
root.geometry("600x400")
root.title('AutoWhats')
root.iconbitmap('icon.ico')
# declaring string variable
# for storing name and password
excel_path_var = tk.StringVar()
gap_time_var = tk.StringVar()
profile_name_var = tk.StringVar()
headless_var = tk.IntVar()
# defining a function that will
# get the name and password and
# print them on the screen
def start():
    excel_path = excel_path_var.get()
    gap_time = gap_time_var.get()
    profile_name = profile_name_var.get()
    # from main import send
    # print("The name is : " + excel_path)
    # print("The password is : " + gap_time)
    send(excel_path=excel_path, gap_break=float(gap_time), profile=profile_name)
    return None
    # excel_path_var.set("")
    # gap_time_var.set("")


# creating a label for
# name using widget Label
excel_path_label = tk.Label(root, text='Excel Path', font=('Bradley Hand ITC', 20, 'bold'))

# creating a entry for input
# name using widget Entry
excel_path_entry = tk.Entry(root, textvariable=excel_path_var, font=('Calibre', 10, 'normal'), width = 45)

# creating a label for password
gap_time_label = tk.Label(root, text='Gap Time', font=('Bradley Hand ITC', 18, 'bold'))

# creating a entry for password
gap_time_entry = tk.Entry(root, textvariable=gap_time_var, font=('calibre', 12, 'normal'))
gap_time_entry.insert(index=0, string=0)
# creating a button using the widget
# Button that will call the start function
sub_btn = tk.Button(root, text='start',font=('Calibre', 10, 'normal'), command=start)

# placing the label and entry in
# the required position using grid
# method
excel_path_label.grid(row=0, column=0)
excel_path_entry.grid(row=0, column=1)
gap_time_label.grid(row=1, column=0)
gap_time_entry .grid(row=1, column=1)
sub_btn.grid(row=2, column=1)







def browsefunc():
    filename = askopenfilename(filetypes=(("xlsx file", "*.xlsx"),("All files", " *.* "),))
    excel_path_entry.delete(0, 'end')
    excel_path_entry.insert(index=0,string=filename)  # add this
    return None
b1=tk.Button(root, text="Browse", font=('Calibre', 10, 'normal'), command=browsefunc)
b1.grid(row=0, column=6)

# Profile:::
separator = ttk.Separator(root, orient='horizontal')
separator.place(relx=0, rely=0.28, relwidth=1, relheight=0.1)
profile_label = tk.Label(root, text='PROFILES:', font=('Ink Free', 20, 'bold'))
profile_label.place(relx=0.1, rely=0.28, relheight=0.3, relwidth=0.8)

profiles = ttk.Combobox(root, width = 27, textvariable = profile_name_var)
# Adding combobox drop down list
with open('setting.txt', 'a', encoding="utf-8") as file:
    file.write('|')
with open('setting.txt', 'r') as file:
    new_sets = (file.read().strip('|').split('|'))
    temp = []
    for new_set in new_sets:
        if new_set != '':
            temp.append(new_set)
    setting = temp
profiles['values'] = setting
profiles.place(relx=0.35, rely=0.5, relheight=0.05, relwidth=0.3)
profiles.current()

def add_profile_popup():
    top = tk.Toplevel(root)
    top.geometry("300x150")

    # Create an Entry Widget in the Toplevel window
    profile_name_entry = tk.Entry(top, width=30,  textvariable=profile_name_var, font=('Calibre', 10, 'normal'))
    profile_name_entry.pack()
    tk.Label(top, text='Enter Profile Name', font=('Bradley Hand ITC', 18, 'bold')).pack()

    # Create a Button Widget in the Toplevel Window
    button = tk.Button(top, text="Ok",font=('Calibre', 10, 'normal'), command=lambda :add_profile(top, profile_name_var))
    button.pack(pady=5, side=tk.TOP)
    return None

def add_profile(window, profile):
    # from main import profile_login
    window.destroy()
    profile = profile.get()
    profile_login(profile)
    with open('setting.txt', 'r') as file:
        new_sets = (file.read().strip('|').split('|'))
        temp = []
        for new_set in new_sets:
            if new_set != '':
                temp.append(new_set)
        setting = temp

    profiles['values'] = setting
    return None

def profile_remove():
    # from main import profile_logout
    profile = profile_name_var.get()
    profile_logout(profile)
    with open('setting.txt', 'r') as file:
        new_sets = (file.read().strip('|').split('|'))
        temp = []
        for new_set in new_sets:
            if new_set != '':
                temp.append(new_set)
        setting = temp
    profiles.delete(0, 'end')
    profiles['values'] = setting
    return None




add_btn=tk.Button(root, text="ADD", font=('Calibre', 10, 'normal'), command=add_profile_popup)
add_btn.place(relx=0.1, rely=0.5, relheight=0.05, relwidth=0.1)


logout_btn=tk.Button(root, text="Log Out", font=('Calibre', 10, 'normal'), command=profile_remove)
logout_btn.place(relx=0.8, rely=0.5, relheight=0.05, relwidth=0.1)


headless_button = tk.Checkbutton(root, text = "Show Browser",
                      variable = headless_var,
                      onvalue = 0,
                      offvalue = 1,
                      height = 2,
                      width = 10,font=('Calibre', 10, 'normal'))

headless_button.place(relx=0.2, rely=0.8, relheight=0.05, relwidth=0.5)








import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys

from datetime import datetime
import shutil
from config import *

user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36"

# Chrome

options = webdriver.ChromeOptions()
# options.headless = True
options.add_argument(f'user-agent={user_agent}')
options.add_argument("--window-size=1920,1080")
options.add_argument('--ignore-certificate-errors')
options.add_argument('--allow-running-insecure-content')
options.add_argument("--disable-extensions")
options.add_argument("--proxy-server='direct://'")
options.add_argument("--proxy-bypass-list=*")
options.add_argument("--start-maximized")
options.add_argument('--disable-gpu')
options.add_argument("--mute-audio")
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')
options.add_argument("--app=https://www.google.com")
options.add_experimental_option("useAutomationExtension", False)
options.add_experimental_option("excludeSwitches", ["enable-automation"])





# options.add_argument("user-data-dir=%s" % (current_profiles_path))
# driver = webdriver.Chrome(executable_path=ChromeDriverManager(log_level=0).install(), options=options)


def open_driver(profile:str, link: str = 'https://web.whatsapp.com', headless: str = True,
                check_login: str = False):
    global driver
    options.headless = headless
    options.add_argument("user-data-dir=%s" % (profiles_path + str(profile)))
    servic = Service(ChromeDriverManager(log_level=0).install())
    driver = webdriver.Chrome(service=servic, options=options)
    driver.maximize_window()
    driver.get(link)
    driver_len = len(driver.window_handles)
    if driver_len > 1:  # Will execute if more than 1 tabs found.
        for i in range(driver_len - 1):
            driver.switch_to.window(driver.window_handles[i])  # will close the last tab first.
            driver.close()
            driver.switch_to.window(driver.window_handles[0])  # Switching the driver focus to First tab.
    time.sleep(0.2)
    if check_login:
        return (check_logined())


def check_logined(wait: int = 300):
    wait = wait // 3  # because it's searching for three elements
    # print('now checking, wait time:', wait)
    if (locate_element(xpath="//div[@title = 'Menu']", wait_time=wait))[0] or \
            locate_element(xpath="//div[@title = 'Status']", wait_time=wait)[0] or \
            locate_element(xpath="//div[@title = 'New chat']", wait_time=wait)[0]:
        return (True, 'profile is actively login')
    else:
        if locate_element(xpath='//div[@class = "_1N3oL"]', wait_time=5)[0]:
            return (False, 'profile already logged out using app')

        else:
            return (False, 'unknown error with profile re-login it again')


def locate_element(xpath: str, wait_time: int = 30, trial_times: int = 10):
    count = 0
    wait_time = wait_time // trial_times
    while count != trial_times:
        try:
            element = WebDriverWait(driver, wait_time).until(EC.presence_of_element_located((By.XPATH, xpath)))
            return True, element
        except:
            count += 1
            # time.sleep(3)
    return False, False


def profile_login(profile_name: str):
    open_driver(profile=profile_name, headless=False)

    if check_logined(3600)[0]:
        time.sleep(1)
        setting_add(profile_name)
        driver.quit()
        return (True, 'added profile ' + str(profile_name))

    else:
        driver.quit()
        return (True, 'error in adding profile, scan again')


def profile_logout(profile: str):
    # open_driver(profile)
    # print('driver opened ')
    check = open_driver(profile,check_login=True, headless = bool(headless_var.get()))
    if check[0]:
        three_dot_btn = locate_element(
            xpath='//div[@title = "Menu"]', wait_time=3,
            trial_times=5)
        if three_dot_btn[0]:
            three_dot_btn[1].click()
            logout_btn = locate_element(
                xpath='//div[@aria-label="Log out"]',
                wait_time=3, trial_times=5)
            if logout_btn[0]:
                time.sleep(1)
                logout_btn[1].click()
                time.sleep(3)
                driver.quit()
                shutil.rmtree(profiles_path + str(profile), ignore_errors=True)

                setting_remove(profile)
                return ('logout successful')


            else:
                driver.quit()
                shutil.rmtree(profiles_path + str(profile), ignore_errors=True)
                return ("error in loging out, menu clicked but can't locate logout button")
        else:
            driver.quit()
            shutil.rmtree(profiles_path + str(profile), ignore_errors=True)
            return ("error in loging out, can't locate menu button")
    else:
        shutil.rmtree(profiles_path + str(profile), ignore_errors=True)
        driver.quit()
        return check[1]





def invalid(num: str = None):
    if num is None:
        invalid_msg = locate_element(
            xpath="//div[@class='_2Nr6U']", wait_time=5)  # web element 'Phone number shared via url is invalid.'
        if invalid_msg[0]:
            return (True, '%sPhone Number is invalid' % num)
        else:
            return (False, False)
    else:
        driver.get('https://web.whatsapp.com/send?phone=' + num)
        invalid_msg = locate_element(
            xpath="//div[@class='_2Nr6U']", wait_time=60)  # web element 'Phone number shared via url is invalid.'
        if invalid_msg[0]:
            return (True, '%sPhone Number is invalid' % num)
        else:
            return (False, False)


def check_sent(): # it will check msg sent or not successfully

    for i in range(2):
        contact = driver.find_element(By.XPATH, '//div[@class= "_2rlF7"]').text
        time.sleep(3)
        chats = driver.find_elements(By.CLASS_NAME, "_3OvU8")
        for chat in chats: #4 coz we can pin 3 chats and one if three are already pinned
            if chat.find_element(By.CLASS_NAME, 'zoWT4').text == contact:
                if locate_element(xpath='//span[@data-testid="status-dblcheck"]', wait_time=15)[0]:
                    return (True, 'Successfully Sent double tick')
                elif locate_element(xpath='//span[@data-testid="status-check"]', wait_time=15)[0]:
                    return (True, 'Successfully Sent, single tick')
                else:
                    return(False, "Msg was sent but couldn't locate single or double tick, Unknown error")
                # elif locate_element(xpath='//span[@data-testid="status-time"]')[0]:
                #     print('watch-clock')
            else:
                if chat == chats[-1]:
                    if i == 1:
                        return (False, " Msg was sent but can't locate it in chats sections, Unknown error")
                    else:
                        time.sleep(3)


def is_emoji(txt: str = None):
    if type(txt)!=str:
        txt = str(txt)
    import re
    lis = re.findall(r'[^\w\s,.!\"#$%&\'()*+\,./:-;<=>?@\[\]^_\\`{|}~]', txt)
    if len(lis) > 0:
        return (True, lis)
    else:
        return (False, lis)
#
def send_txt(num: str = None, msg: str = None):
    if msg is not None:
        emoji = is_emoji(msg)
    else:
        return (False, "no msg found")

    if num is not None:
        # num = str(num)
        if num.startswith("+"):
            from urllib.parse import quote
            try:
                parsed_message = quote(msg)
            except:
                parsed_message = str(msg)

            driver.get('https://web.whatsapp.com/send?phone=' +
                       num + '&text=' + parsed_message)
            # time.sleep(1)

        else:
            return (False, "missing country code", num)
    else:
        pass

    send_box = locate_element(xpath='//div[@title="Type a message"]', wait_time=3)

    if send_box[0]:
        if num is not None:
            send_box[1].click()
            ActionChains(driver).send_keys(Keys.ENTER).perform()
        else:
            if emoji[0]:
                send_box[1].send_keys(msg)
                # time.sleep(0.2)
                ActionChains(driver).send_keys(Keys.ENTER).perform()
                # import pyperclip
                # past = pyperclip.paste()
                # pyperclip.copy(msg)
                # send_box[1].click()
                # action = ActionChains(driver)
                # action.key_down(Keys.LEFT_CONTROL).send_keys('v').key_up(Keys.LEFT_CONTROL).perform()
                # pyperclip.copy(past)
                # time.sleep(0.2)
                # action.send_keys(Keys.ENTER).perform()
            else:
                send_box[1].send_keys(msg)
                # time.sleep(0.2)
                ActionChains(driver).send_keys(Keys.ENTER).perform()
                # import pyperclip
                # past = pyperclip.paste()
                # pyperclip.copy(msg)
                # send_box[1].click()
                # action = ActionChains(driver)
                # action.key_down(Keys.LEFT_CONTROL).send_keys('v').key_up(Keys.LEFT_CONTROL).perform()
                # pyperclip.copy(past)
                # time.sleep(0.2)
                # action.send_keys(Keys.ENTER).perform()

        try:
            send_box[1].clear()
        except:
            pass
        #  checking sent or not:
        status = check_sent()
        if status[0]:
            return(True, status[1])
        else:
            return (False, status[1])
    else:
        if num is not None:
            if invalid()[0]:
                return (False, 'Phone Number: %s is not active in WhatsApp'%num)
        else:
            if check_logined()[0]:
                return (False, "Unknown error in sending")
            else:
                return (False, "Profile is missing or already logged out")






def send_media(num: str = None, loc: str = None, caption: str = None):
    if loc is not None:
        if num is not None:
            # num = str(num)
            if num.startswith("+"):
                if caption is not None:
                    from urllib.parse import quote
                    try:
                        parsed_message = quote(caption)
                    except:
                        parsed_message = str(caption)
                    driver.get('https://web.whatsapp.com/send?phone=' +
                               num + '&text=' + parsed_message)
                else:
                    driver.get('https://web.whatsapp.com/send?phone=' + num)
                # time.sleep(1)

            else:
                return (False, "missing country code", num)
        else:
            if caption is not None:
                send_box = locate_element(xpath='//div[@title="Type a message"]', wait_time=3)
                if send_box[0]:
                    emoji = is_emoji(caption)
                    if emoji[0]:
                        import pyperclip
                        past = pyperclip.paste()
                        pyperclip.copy(caption)
                        action = ActionChains(driver)
                        action.key_down(Keys.LEFT_CONTROL).send_keys('v').key_up(Keys.LEFT_CONTROL).perform()
                        pyperclip.copy(past)
                        time.sleep(0.2)
                    else:
                        send_box[1].send_keys(caption)
                        time.sleep(0.2)
                else:
                    if num is not None:
                        if invalid()[0]:
                            return (False, 'Phone Number: %s is not active in WhatsApp' % num)
                    else:
                        pass
                    if check_logined()[0]:
                        return (False, "couldn't locate send box")
                    else:
                        return (False, "Profile is missing or already logged out")
            else:
                pass
        attach_clip = locate_element(xpath='//span[@data-testid="clip"]', wait_time=3)

        if attach_clip[0]:
            attach_clip[1].click()
            media_btn = locate_element(xpath='//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]', wait_time=3)
            if media_btn[0]:
                try:
                    media_btn[1].send_keys(loc)
                except:
                    return (False, "error in locating file or either file too large")

                if locate_element(xpath='//div[@class="_7Tfhi"]', wait_time=20)[0]:
                    ActionChains(driver).send_keys(Keys.ENTER).perform()
                    #  checking sent or not:
                    status = check_sent()
                    if status[0]:
                        return (True, status[1])
                    else:
                        return (False, status[1])
                else:
                    return (False, "couldn't locate file after uploading")
            else:
                if num is None:
                    return (False, "couldn't locate the img/vid button")
                else:
                    if num is not None:
                        if invalid()[0]:
                            return (False, 'Phone Number: %s is not active in WhatsApp' % num)
                    else:
                        pass
                    if check_logined()[0]:
                        return (False, "couldn't locate send box")
                    else:
                        return (False, "Profile is missing or already logged out")



        else:
            if num is not None:
                if invalid()[0]:
                    return (False, 'Phone Number: %s is not active in WhatsApp' % num)
            else:
                if check_logined()[0]:
                    return (False, "Unknown error in sending")
                else:
                    return (False, "Profile is missing or already logged out")

    else:
        return(False, "file path can't be none")


def send_doc(num: str = None, loc: str = None):
    if loc is not None:
        if num is not None:
            # num = str(num)
            if num.startswith("+"):
                driver.get('https://web.whatsapp.com/send?phone=' + num)
                # time.sleep(1)

            else:
                return (False, "missing country code", num)
        else:
            pass
        attach_clip = locate_element(xpath='//span[@data-testid="clip"]', wait_time=3)

        if attach_clip[0]:
            attach_clip[1].click()
            doc_btn = locate_element(xpath='//input[@accept="*"]', wait_time=3)
            if doc_btn[0]:
                try:
                    doc_btn[1].send_keys(loc)
                except:
                    return (False, "error in locating file or either file too large")
                if locate_element(xpath='//div[@class="_7Tfhi"]', wait_time=20)[0]:
                    ActionChains(driver).send_keys(Keys.ENTER).perform()
                    time.sleep(2)
                    #  checking sent or not:
                    status = check_sent()
                    if status[0]:
                        return (True, status[1])
                    else:
                        return (False, status[1])
                else:
                    return (False, "couldn't locate file after uploading")


            else:
                return (False, "couldn't locate documents button")


        else:
            if num is not None:
                if invalid()[0]:
                    return (False, 'Phone Number: %s is not active in WhatsApp' % num)
            else:
                pass
            if check_logined()[0]:
                return (False, "couldn't locate attach clip")
            else:
                return (False, "Profile is missing or already logged out")

    else:
        return(False, "file path can't be none")


def send_sticker(num: str = None, loc: str = None):
    if loc is not None:
        if num is not None:
            # num = str(num)
            if num.startswith("+"):
                driver.get('https://web.whatsapp.com/send?phone=' + num)
                # time.sleep(1)

            else:
                return (False, "missing country code", num)
        else:
            pass
        attach_clip = locate_element(xpath='//span[@data-testid="clip"]', wait_time=3)
        if attach_clip[0]:

            attach_clip[1].click()
            sticker_btn = locate_element(xpath='//input[@accept="image/*"]', wait_time=3)
            if sticker_btn[0]:
                try:
                    sticker_btn[1].send_keys(loc)
                except:
                    return (False, "error in locating file or either file too large")
                if locate_element(xpath='//div[@class="_7Tfhi"]', wait_time=20)[0]:
                    ActionChains(driver).send_keys(Keys.ENTER).perform()
                    # time.sleep(2)
                    try:
                        WebDriverWait(driver, 15).until(
                            EC.element_to_be_clickable((By.XPATH, '//span[@data-testid="clip"]')))
                        time.sleep(8)
                    except:
                        return (False, 'error in finding chat column after sending')
                    #  checking sent or not:
                    status = check_sent()
                    if status[0]:
                        return (True, status[1])
                    else:
                        return (False, status[1])
                else:
                    return (False, "couldn't locate file after uploading")

            else:
                return (False, "couldn't locate sticker button")


        else:
            if num is not None:
                if invalid()[0]:
                    return (False, 'Phone Number: %s is not active in WhatsApp' % num)
            else:
                pass
            if check_logined()[0]:
                return (False, "couldn't locate attach clip")
            else:
                return (False, "Profile is missing or already logged out")

    else:
        return (False, "file path can't be none")





# Send::::

def send(profile:str = None, excel_path: str = None, gap_set:int = 1,
         gap_break:int= 0.5, history: str = False,
         report: str = True, send_report: str = False,
         send_report_to: str = None, send_xl_report:str= False):


    from plyer import notification
    import openpyxl
    from openpyxl.styles import PatternFill

    if open_driver(profile=profile, check_login=True, headless = bool(headless_var.get()))[0]:

        sheet_num = 1


        whole_sent_count = 0

        wb = openpyxl.load_workbook(excel_path)
        excel = wb[wb.sheetnames[sheet_num-1]]
        total_row = excel.max_row

        for track_row in range(2,total_row+1):
            num = None
            msg = None
            media = None
            sticker = None
            doc = None
            total_to_send = 0
            thing_sent = 0
            status = []
            print('sending....' + str(track_row-1) + ' / ' + str(total_row-1))
            if track_row == 2:
                notification.notify(
                    title="Started sending",
                    message="Sending Started",
                    app_name='AutoWhats',
                    app_icon='icon.ico',
                    # displaying time
                    timeout=3
                )

            num = excel.cell(track_row, 1).value #clmnA= 'contact'
            try:
                excel.cell(track_row, 4).value = (excel.cell(track_row, 2).value) % tuple(
                    (excel.cell(track_row, 3).value).split(','))
                msg = excel.cell(track_row, 4).value
                wb.save(excel_path.replace('.xlsx','_Report.xlsx'))
            except AttributeError:
                excel.cell(track_row, 4).value = excel.cell(track_row, 2).value
                msg = excel.cell(track_row, 4).value
                wb.save(excel_path.replace('.xlsx','_Report.xlsx'))

            except TypeError:
                excel.cell(track_row, 4).value = '%s and values given to it is mismatched'
                excel.cell(track_row, 4).fill = PatternFill('solid', fgColor='FF0000')

                wb.save(excel_path.replace('.xlsx','_Report.xlsx'))
                thing_sent -= 1
                status.append('text msg: error in insertion')
            finally:
                wb.save(excel_path.replace('.xlsx','_Report.xlsx'))

            media = excel.cell(track_row, 5).value
            sticker = excel.cell(track_row, 6).value
            doc = excel.cell(track_row, 7).value

            if num:
                driver.get('https://web.whatsapp.com/send?phone=' + num)
                send_box = locate_element(xpath='//div[@title="Type a message"]', wait_time=60)

                if send_box[0]:
                    if msg:
                        total_to_send += 1
                        sending= send_txt(msg=msg)
                        status.append('text msg:' + sending[1])
                        if sending[0]:
                            thing_sent += 1
                        else:
                            excel.cell(track_row, 4).fill = PatternFill('solid', 'FF0000')

                    else:
                        pass

                    if media:
                        list_media = media.split(',')
                        for one_media in list_media:
                            total_to_send += 1
                            sending = send_media(loc=one_media.strip())
                            status.append('media:' + sending[1])
                            if sending[0]:
                                thing_sent += 1
                            else:
                                excel.cell(track_row, 5).fill = PatternFill('solid', 'FF0000')

                    else:
                        pass

                    if sticker:
                        list_sticker = sticker.split(',')
                        for one_sticker in list_sticker:
                            total_to_send += 1
                            sending = send_sticker(loc=one_sticker.strip())
                            status.append('sticker:' + sending[1])
                            if sending[0]:
                                thing_sent += 1
                            else:
                                excel.cell(track_row, 6).fill = PatternFill('solid', 'FF0000')

                    else:
                        pass


                    if doc:
                        list_doc = doc.split(',')
                        for one_doc in list_doc:
                            total_to_send += 1
                            sending = send_doc(loc=one_doc.strip())
                            status.append('doc:' + sending[1])
                            if sending[0]:
                                thing_sent += 1
                            else:
                                excel.cell(track_row, 7).fill = PatternFill('solid', 'FF0000')

                    else:
                        pass

                else:
                    if invalid()[0]:
                        status.append('Invalid Number')
                        excel.cell(track_row, 8).fill = PatternFill('solid', 'FF0000')
                        wb.save(excel_path.replace('.xlsx','_Report.xlsx'))
                        total_to_send -= 1
                    elif check_logined()[0]:
                        status.append('Unknown error in sending')
                        excel.cell(track_row, 8).fill = PatternFill('solid', 'FF0000')
                        wb.save(excel_path.replace('.xlsx','_Report.xlsx'))
                        total_to_send -= 1
                    else:
                        status.append('WhatsApp profile missing or logout, try re-login')
                        excel.cell(track_row, 8).fill = PatternFill('solid', 'FF0000')
                        wb.save(excel_path.replace('.xlsx','_Report.xlsx'))
                        total_to_send -= 1

                excel.cell(track_row, 8).value = str(status)
                if total_to_send == thing_sent:
                    if thing_sent == 0:
                        excel.cell(track_row, 8).fill = PatternFill('solid', 'FF0000')
                        wb.save(excel_path.replace('.xlsx','_Report.xlsx'))

                    else:
                        excel.cell(track_row, 8).fill = PatternFill('solid', '00FF00')
                        wb.save(excel_path.replace('.xlsx','_Report.xlsx'))
                        whole_sent_count += 1

                elif total_to_send == 0:
                    excel.cell(track_row, 8).value = 'nothing to send'
                    excel.cell(track_row, 8).fill = PatternFill('solid', 'FF0000')
                    wb.save(excel_path.replace('.xlsx','_Report.xlsx'))


                else:
                    excel.cell(track_row, 8).fill = PatternFill('solid', 'FF0000')
                    wb.save(excel_path.replace('.xlsx','_Report.xlsx'))

            if (track_row-1)%(gap_set) == 0:
                time.sleep(gap_break)

            # sierry
            # with open('number.txt', 'r') as file:
            #     number = int(file.read())
            #     if number + track_row > 619:
            #         import random
            #         if track_row > 30 and bool(random.randint(0,1)):
            #             track_row == total_row
            #
            #         if track_row == track_row - random.randrange(0,20):
            #             driver.quit()
            #             break


            if track_row == total_row:
                # shutil.move('Report.xlsx', excel_path.replace('.xlsx','_Report.xlsx'))
                driver.quit()
                notification.notify(
                    title="Sent Messages Sucessfully",
                    message=f"Total {whole_sent_count} Messages are sent out of {total_row-1} msgs, you can see the report {excel_path.replace('.xlsx','_Report.xlsx')}.",
                    app_name= 'AutoWhats',
                    app_icon ='icon.ico',
                    # displaying time
                    timeout=7
                )


                 # seeriy
                # with open('number.txt', 'w') as file:
                #     file.write(str(number + whole_sent_count))

                # if 'yes' in send_report_to_WhatsApp:
                #     if 'yes' in shutdwn:
                #         sendwhatmsg_instantly_txt(str(send_num), f'Out of {total_row-1} messages Total {count} msgs sent. Your device is about to turn off after 60 secs')
                #     else:
                #         sendwhatmsg_instantly_txt(str(send_num), f'Out of {total_row-1} messages Total {count} msgs sent.')
                # sleep(3)

                # if 'yes' in shutdwn:
                #     shutdown()


# send(excel_path='C:\\Users\\PRINCE MEHTA\\Desktop\\2.xlsx')
# setting:::
def setting_add(name):
    with open('setting.txt','a', encoding="utf-8") as file:
       file.write(str(name)+'|')

def setting_remove(name):
   with open('setting.txt', 'r') as file:
      new_sets = (file.read().strip('|').split('|'))
      temp = []
      for new_set in new_sets:
        if new_set != '':
            temp.append(new_set)
      new_setting = temp
   new_setting.remove(name)
   with open('setting.txt', 'w', encoding="utf-8") as file:
      file.write('')
   for set in new_setting:
      setting_add(set)



# performing an infinite loop
# for the window to display
root.mainloop()